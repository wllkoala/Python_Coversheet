import os
import traceback
from sys import exit
from tkinter import Tk, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from PyPDF4 import PdfFileReader, PdfFileWriter
from win32com.client import DispatchEx

root = Tk()
root.withdraw()


class AddCoverSheet():
    name_lists = []
    doc_codes = []
    doc_revs = []
    file_lists = []

    def __init__(self, file_dir):
        self.file_dir = file_dir

    def start_to_run(self):
        '''开始运行'''
        self.create_folder()
        self.get_name_lists()
        self.cover_sheet()
        self.conversion()
        self.merge_doc()
        messagebox.showinfo("Complete!", "全部文件已完成！")

    def create_folder(self):
        '''创建所需文件夹'''
        if not os.path.exists("input"):
            os.mkdir("input")
        if not os.path.exists("output"):
            os.mkdir("output")
        else:
            self.del_file("output")
        if not os.path.exists("tmp"):
            os.mkdir("tmp")
        else:
            self.del_file("tmp")

    def del_file(self, folder):
        '''删除已有文件'''
        for file in os.listdir(folder):
            path_file = os.path.join(folder, file)
            if os.path.isfile(path_file):
                os.remove(path_file)

    def get_name_lists(self):
        '''获取待添加封面文件目录'''
        for name in os.listdir("input"):
            if name.endswith(".pdf"):
                if name.find('_') > 0:
                    self.doc_codes.append(name.split("_")[0])
                    self.doc_revs.append(name.split("_")[1][1:3])
                    self.name_lists.append(name)
                    print("需要添加封面的文件：", len(self.name_lists),
                          name.split(".")[0])
        print("=><=" * 25)

    def conversion(self):
        '''转换封面EXCEL为PDF'''
        xlApp = DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = 0
        for name_list_index, name_list in enumerate(self.doc_codes):
            print('当前文件转换进度',
                  name_list_index + 1, "/", len(self.doc_codes))
            exportfile = name_list
            filenames = exportfile.split('.')[0] + '.xlsx'
            filename = filenames.replace("input", "tmp")
            books = xlApp.Workbooks.Open(filename, False)
            books.ExportAsFixedFormat(0, exportfile)
            books.Close(False)
            print('封面转为PDF文件：', exportfile)
        xlApp.Quit()
        print('封面转为PDF文件完成')
        print("=><=" * 25)

    def merge_doc(self):
        '''合并封面和文件'''
        self.file_lists = list(zip(self.doc_codes, self.name_lists))
        for pdfnames in self.file_lists:
            output = PdfFileWriter()
            for pdfname in pdfnames:
                input = PdfFileReader(open(pdfname, "rb"), strict=False)
                pageCount = input.getNumPages()
                for iPage in range(0, pageCount):
                    output.addPage(input.getPage(iPage))
            pdfoutname = str(pdfnames[0]).replace("input", "output")
            outputStream = open(pdfoutname, "wb")
            output.write(outputStream)
            outputStream.close()
            print("文件合并完成：", pdfoutname)
        print("文件合并完成！")
        print("=><=" * 25)

    def cover_sheet(self):
        '''生成excel版封面'''
        excel_file = filedialog.askopenfilename(
            title="Select the file", filetypes=[("All files", "*")])
        df = pd.read_excel(excel_file)
        df = df.dropna(axis=0, how='all')
        df = df.fillna("NA")
        df = df.reset_index(drop=True)
        col_names = df.columns.values.tolist()
        print("需要生成封面文件数：", len(self.doc_codes))
        num = []
        for n, doc_code in enumerate(self.doc_codes):
            if doc_code in df.iloc[:, 0].values:
                doc_code_index = df[df.iloc[:, 0] == doc_code].index.tolist()
                doc_data = df.iloc[doc_code_index].values[0]
                wb = load_workbook("Coversheet.xlsx")
                ws = wb["Tempdata"]
                for i, col_name in enumerate(col_names):
                    ws.cell(i+1, 1).value = col_name
                    ws.cell(i+1, 2).value = doc_data[i]
                file_name = os.path.join("tmp", doc_code + ".xlsx")
                wb.save(file_name)
                print("当前封面生成进度：", n + 1, "/", len(self.doc_codes))
                print("文件封面已完成", doc_code)
                self.doc_codes[n] = os.path.join(
                    self.file_dir, "input", doc_code + '.pdf')
                self.name_lists[n] = os.path.join(
                    self.file_dir, "input", self.name_lists[n])
            else:
                num.append(n)
                print("当前封面生成进度：", n + 1, "/", len(self.doc_codes))
                print("文件信息未找到", doc_code)
                continue
        for i, n in enumerate(num):
            del self.doc_codes[n-i]
            del self.doc_revs[n-i]
            del self.name_lists[n-i]
        print("=><=" * 25)


if __name__ == "__main__":
    try:
        file_dir = os.getcwd()
        print("=><=" * 25)
        print("当前路径：", file_dir)
        print("=><=" * 25)
        merge = AddCoverSheet(file_dir)
        merge.start_to_run()
    except Exception as err:
        messagebox.showerror("Warning!", err)
        with open(os.path.join(os.getcwd(), "error.txt"), "w") as f:
            traceback.print_exc(file=f)
        print(err)
        exit()
